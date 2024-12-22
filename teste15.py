import customtkinter
import tkinter as tk
from tkinter import messagebox, filedialog
from datetime import datetime
from openpyxl import Workbook
import os
from PIL import Image, ImageTk
import pandas as pd
from tkinter import Tk, Label, Entry, Button, messagebox
from tkinter import *
from tkinter import ttk
from PIL import Image, ImageTk


customtkinter.set_appearance_mode("dark")
customtkinter.set_default_color_theme("dark-blue")

class ManutencaoApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Manutenção Hoteleira")

        # Criando os rótulos e campos de entrada
        self.create_widgets()

        # Caminho do arquivo Excel
        self.file_path = r"C:\Users\thays\OneDrive\Área de Trabalho\Aplicativo_de_Manutencao\Ordem de Serviço.xlsx"

        # Verifica se o arquivo já existe, se não, cria um novo
        if not os.path.exists(self.file_path):
            self.create_excel_file()

    def create_widgets(self):
        # Campo Data
        tk.Label(self.root, text='Data do Atendimento').grid(row=0, column=0)
        self.data_entry = tk.Entry(self.root)
        self.data_entry.grid(row=0, column=1)
        
        # Preencher automaticamente o campo de data com a data atual
        self.data_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))

        # Campo Local do Atendimento
        tk.Label(self.root, text='Local').grid(row=1, column=0)
        self.local_entry = tk.Entry(self.root)
        self.local_entry.grid(row=1, column=1)

        # Campo Oficina
        tk.Label(self.root, text='Oficina').grid(row=2, column=0)
        self.oficina_entry = tk.Entry(self.root)
        self.oficina_entry.grid(row=2, column=1)

        # Campo Serviço
        tk.Label(self.root, text='Descrição do Serviço').grid(row=3, column=0)
        self.servico_entry = tk.Entry(self.root)
        self.servico_entry.grid(row=3, column=1)

        # Campo Material Utilizado
        tk.Label(self.root, text='Material Utilizado').grid(row=4, column=0)
        self.material_entry = tk.Entry(self.root)
        self.material_entry.grid(row=4, column=1)
        
        # Campo Quantidade
        tk.Label(self.root, text='Quantidade de Material').grid(row=5, column=0)
        self.quantidade_entry = tk.Entry(self.root)
        self.quantidade_entry.grid(row=5, column=1)

        # Campo Status do Atendimento
        tk.Label(self.root, text='Status do Atendimento').grid(row=6, column=0)
        
        # Variável para armazenar a seleção
        self.status_var = tk.StringVar(value='Aberto')  # Valor padrão
        
        # Opções para o Status
        self.status_option = tk.OptionMenu(self.root, self.status_var, 'Concluído', 'Aberto', command=self.update_status_color)
        self.status_option.grid(row=6, column=1)

        # Campo Observação
        tk.Label(self.root, text='Observação').grid(row=7, column=0)
        self.observacao_entry = tk.Text(self.root, height=4, width=30)
        self.observacao_entry.grid(row=7, column=1)

        # Campo Anexo
        tk.Label(self.root, text='Anexo (opcional)').grid(row=8, column=0)
        self.anexo_entry = tk.Entry(self.root)
        self.anexo_entry.grid(row=8, column=1)
        self.anexo_button = tk.Button(self.root, text='Selecionar Arquivo', command=self.select_file)
        self.anexo_button.grid(row=8, column=2)

        # Botão para enviar os dados
        self.submit_button = tk.Button(self.root, text='Enviar', command=self.submit_form)
        self.submit_button.grid(row=9, column=0, columnspan=3)

        # Inicializa a cor do campo baseado na seleção padrão
        self.update_status_color(self.status_var.get())

    def create_excel_file(self):
        # Cria um novo arquivo Excel e adiciona o cabeçalho
        wb = Workbook()
        ws = wb.active
        ws.title = "Ordens de Serviço"
        
        # Adiciona cabeçalhos
        headers = ['Data', 'Local', 'Oficina', 'Descrição do Serviço', 'Material Utilizado', 'Quantidade de Material', 'Status do Atendimento', 'Observação', 'Anexo']
        ws.append(headers)
        
        wb.save(self.file_path)

    def submit_form(self):
        # Coletar dados
        data = self.data_entry.get()
        local = self.local_entry.get()
        oficina = self.oficina_entry.get()
        servico = self.servico_entry.get()
        material = self.material_entry.get()
        quantidade = self.quantidade_entry.get()
        status = self.status_var.get()
        observacao = self.observacao_entry.get("1.0", tk.END).strip()
        anexo = self.anexo_entry.get()

        # Adiciona os dados à planilha
        self.add_data_to_excel([data, local, oficina, servico, material, quantidade, status, observacao, anexo])

        # Exibir mensagem de sucesso
        messagebox.showinfo("Sucesso", "Dados enviados com sucesso!")

        # Limpar campos após envio
        self.clear_fields()

    def add_data_to_excel(self, data):
        from openpyxl import load_workbook
        
        # Carrega o arquivo existente
        wb = load_workbook(self.file_path)
        ws = wb.active
        
        # Adiciona uma nova linha com os dados
        ws.append(data)
        
        # Salva o arquivo
        wb.save(self.file_path)

    def clear_fields(self):
        self.data_entry.delete(0, tk.END)
        self.data_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))  # Repreencher com a data atual
        self.local_entry.delete(0, tk.END)
        self.oficina_entry.delete(0, tk.END)
        self.servico_entry.delete(0, tk.END)
        self.material_entry.delete(0, tk.END)
        self.quantidade_entry.delete(0, tk.END)
        self.status_var.set('Aberto')  # Resetar para o valor padrão
        self.update_status_color(self.status_var.get())  # Atualiza a cor
        self.observacao_entry.delete("1.0", tk.END)
        self.anexo_entry.delete(0, tk.END)  # Limpa o campo de anexo

    def update_status_color(self, status):
        # Altera a cor do campo com base na seleção do status
        if status == 'Concluído':
            self.status_option.config(bg='green', fg='white')
        else:
            self.status_option.config(bg='red', fg='white')

    def select_file(self):
        # Abre um diálogo para selecionar um arquivo
        file_path = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf"), ("PNG Files", "*.png"), ("JPEG Files", "*.jpg;*.jpeg")])
        if file_path:
            self.anexo_entry.delete(0, tk.END)  # Limpa o campo anterior
            self.anexo_entry.insert(0, file_path)  # Insere o caminho do arquivo selecionado
            
class LoginApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Login")

        # Carregar a imagem
        self.load_image()

        self.texto = customtkinter.CTkLabel(self.master, text="Fazer Login")
        self.texto.pack(padx=8, pady=8)

        self.email = customtkinter.CTkEntry(self.master, placeholder_text="Seu e-mail")
        self.email.pack(padx=10, pady=10)

        self.senha = customtkinter.CTkEntry(self.master, placeholder_text="Sua senha", show="*")
        self.senha.pack(padx=10, pady=10)

        self.checkbox = customtkinter.CTkCheckBox(self.master, text="Lembrar Login")
        self.checkbox.pack(padx=10, pady=10)

        self.botao_login = customtkinter.CTkButton(self.master, text="Login", command=self.fazer_login)
        self.botao_login.pack(padx=10, pady=10)

        self.botao_cadastro = customtkinter.CTkButton(self.master, text="Cadastre-se", command=self.abrir_cadastro)
        self.botao_cadastro.pack(padx=10, pady=10)

    def load_image(self):
        # Carregue a imagem
        self.img = Image.open(r"C:\Users\thays\OneDrive\Área de Trabalho\Aplicativo_de_Manutencao\Arte1_login.png") 
        self.img = self.img.resize((100, 100), Image.LANCZOS)
        self.photo = ImageTk.PhotoImage(self.img)

        # Adiciona a imagem ao rótulo
        self.img_label = tk.Label(self.master, image=self.photo)
        self.img_label.pack(pady=10)  

    def fazer_login(self):
        # Aqui você pode adicionar sua lógica de autenticação
        email_usuario = self.email.get()
        senha_usuario = self.senha.get()

        # Simulação de login bem-sucedido
        if email_usuario and senha_usuario:  # Verifique se os campos não estão vazios
            self.master.destroy()  # Fecha a janela de login
            root = tk.Tk()
            app = ManutencaoApp(root)
            root.mainloop()
        else:
            messagebox.showerror("Erro", "Por favor, preencha todos os campos.")

    def abrir_cadastro(self):
        cadastro_janela = tk.Toplevel(self.master)
        CadastroApp(cadastro_janela)

class CadastroApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Cadastro de Usuário")

        # Criação dos elementos da interface gráfica
        Label(master, text="Nome:").grid(row=0)
        Label(master, text="E-mail:").grid(row=1)
        Label(master, text="Senha:").grid(row=2)
        Label(master, text="Setor:").grid(row=3)

        self.nome_entry = Entry(master)
        self.email_entry = Entry(master)
        self.senha_entry = Entry(master)

        self.nome_entry.grid(row=0, column=1)
        self.email_entry.grid(row=1, column=1)
        self.senha_entry.grid(row=2, column=1)

        # Carregar os setores do arquivo Excel
        self.setores = self.carregar_setores()
        self.setor_combobox = ttk.Combobox(master, values=self.setores)
        self.setor_combobox.grid(row=3, column=1)

        self.cadastrar_button = Button(master, text="Cadastrar", command=self.finalizar_cadastro)
        self.cadastrar_button.grid(row=4, columnspan=2)

    def carregar_setores(self):
        try:
            df_setores = pd.read_excel(r"C:\Users\thays\OneDrive\Área de Trabalho\Aplicativo_de_Manutencao\cadastro_setor.xlsx")
            return df_setores['Setor'].tolist()  # Supondo que a coluna se chama 'Setor'
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível carregar os setores: {e}")
            return []

    def salvar_cadastro(self, nome, email, senha, setor):
        try:
            # Carregar dados existentes, se houver
            try:
                df = pd.read_excel(r"C:\Users\thays\OneDrive\Área de Trabalho\Aplicativo_de_Manutencao\cadastro_usuario.xlsx")
            except FileNotFoundError:
                df = pd.DataFrame(columns=["Nome", "E-mail", "Senha", "Setor"])

            novo_registro = pd.DataFrame({"Nome": [nome], "E-mail": [email], "Senha": [senha], "Setor": [setor]})

            df = pd.concat([df, novo_registro], ignore_index=True)

            df.to_excel(r"C:\Users\thays\OneDrive\Área de Trabalho\Aplicativo_de_Manutencao\cadastro_usuario.xlsx", index=False)

            messagebox.showinfo("Sucesso", "Cadastro realizado com sucesso!")

        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível salvar os dados: {e}")

    def finalizar_cadastro(self):
        nome_usuario = self.nome_entry.get()
        email_usuario = self.email_entry.get()
        senha_usuario = self.senha_entry.get()
        setor_usuario = self.setor_combobox.get()

        if nome_usuario and email_usuario and senha_usuario and setor_usuario:  # Verifica se todos os campos estão preenchidos
            self.salvar_cadastro(nome_usuario, email_usuario, senha_usuario, setor_usuario)
        else:
            messagebox.showwarning("Atenção", "Por favor, preencha todos os campos.")
            
if __name__ == "__main__":
    login_janela = customtkinter.CTk()
    login_app = LoginApp(login_janela)
    login_janela.geometry("500x500")
    login_janela.mainloop()